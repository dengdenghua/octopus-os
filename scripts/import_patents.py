#!/usr/bin/env python3
"""Import a patent corpus (XLSX/CSV) into a company project.

Usage:
    python scripts/import_patents.py <project_id> <path_to_xlsx_or_csv>

Example:
    python scripts/import_patents.py proj_abc123 "D:/Patents/Eight Sleep 爱伊特睡眠相关专利列表-127件有效专利.XLSX"

This script implements the column-detection and dedup logic described in the
patent-import-xlsx skill (see runtime/sensing/siphon/agent_market_sources/
hardware-startup/agent-plugins/patent-fto-screener/skills/patent-import-xlsx/).

Output: a summary report (created/updated/skipped counts) and saves to the
project's company_projects.json via the CompanyStore + bulk-import API.
"""
from __future__ import annotations

import sys
from pathlib import Path

# Adjust sys.path to import from runtime
_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from runtime.company.core import (  # noqa: E402
    BulkImportPatentRecordsRequest,
    CompanyStore,
    CreatePatentRecordRequest,
)


def main() -> None:
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    project_id = sys.argv[1]
    file_path = Path(sys.argv[2])

    if not file_path.exists():
        print(f"Error: file not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    store = CompanyStore()
    project = store.get_project(project_id)
    if project is None:
        print(f"Error: project not found: {project_id}", file=sys.stderr)
        sys.exit(1)

    print(f"Importing {file_path.name} into project {project.name} ({project_id})...")

    # Parse the file
    if file_path.suffix.lower() in (".xlsx", ".xls"):
        records = _parse_xlsx(file_path)
    elif file_path.suffix.lower() == ".csv":
        records = _parse_csv(file_path)
    else:
        print(f"Error: unsupported file type: {file_path.suffix}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsed {len(records)} record(s) from {file_path.name}")

    # Bulk import
    body = BulkImportPatentRecordsRequest(
        records=records,
        source_label=file_path.name,
    )
    counts = store.bulk_import_patents(project_id, body)
    if counts is None:
        print(f"Error: project {project_id} not found during import", file=sys.stderr)
        sys.exit(1)

    print("\nImport complete:")
    print(f"  Created: {counts['created']}")
    print(f"  Updated: {counts['updated']}")
    print(f"  Skipped: {counts['skipped']}")
    print(f"  Total:   {counts['created'] + counts['updated'] + counts['skipped']}")


def _parse_xlsx(path: Path) -> list[CreatePatentRecordRequest]:
    """Parse an XLSX/XLS file into patent records.

    Implements the column-alias table from the patent-import-xlsx skill.
    """
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active

    # Read header row
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = _detect_columns(header_row)

    print(f"Detected columns: {list(col_map.keys())}")

    records: list[CreatePatentRecordRequest] = []
    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        record = _row_to_patent_record(col_map, row_values)
        if record is not None:
            records.append(record)

    return records


def _parse_csv(path: Path) -> list[CreatePatentRecordRequest]:
    """Parse a CSV file into patent records."""
    import csv

    with path.open(encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header_row = next(reader)
        col_map = _detect_columns(header_row)

        print(f"Detected columns: {list(col_map.keys())}")

        records: list[CreatePatentRecordRequest] = []
        for row_values in reader:
            record = _row_to_patent_record(col_map, row_values)
            if record is not None:
                records.append(record)

    return records


def _detect_columns(header_row: tuple[str, ...] | list[str]) -> dict[str, int]:
    """Map logical field names to column indices based on header aliases.

    Implements the column-alias table from the patent-import-xlsx skill.
    """
    ALIASES = {  # noqa: N806
        "title": ["标题", "名称", "发明名称", "专利名称", "title", "patent title", "invention title"],
        "applicant": ["申请人", "专利权人", "申请单位", "[标]当前申请", "当前申请", "申请(专利权)人", "applicant", "assignee", "patent owner"],
        "publicationNumber": ["公开号", "公告号", "公开(公告)号", "publication number", "publication no", "pub no"],
        "applicationNumber": ["申请号", "application number", "app no"],
        "country": ["国家", "国别", "申请国家", "受理局", "country", "jurisdiction", "office"],
        "publicationDate": ["公开日", "公告日", "公开(公告)日", "publication date", "pub date"],
        "filingDate": ["申请日", "filing date"],
        "legalStatus": ["法律状态", "法律状态/事件", "当前状态", "legal status", "status"],
        "abstract": ["摘要", "摘要(译)", "简介", "abstract", "summary"],
        "keyClaimsSummary": ["主权利要求", "权利要求摘要", "independent claim", "claim summary"],
        "inventors": ["发明人", "inventor", "inventors"],
        "ipcCodes": ["ipc", "ipc 分类号", "ipc分类号", "ipc class"],
        "url": ["链接", "url", "link"],
    }

    col_map: dict[str, int] = {}
    for col_idx, header_cell in enumerate(header_row):
        if header_cell is None:
            continue
        header_norm = str(header_cell).strip().lower()
        for logical_field, aliases in ALIASES.items():
            if any(alias in header_norm for alias in aliases):
                col_map[logical_field] = col_idx
                break

    return col_map


def _row_to_patent_record(
    col_map: dict[str, int],
    row_values: tuple[str, ...] | list[str],
) -> CreatePatentRecordRequest | None:
    """Convert a row into a CreatePatentRecordRequest.

    Implements the coercion rules from the patent-import-xlsx skill.
    """
    def _get(field: str) -> str:
        idx = col_map.get(field)
        if idx is None or idx >= len(row_values):
            return ""
        val = row_values[idx]
        if val is None:
            return ""
        return str(val).strip()

    title = _get("title")
    if not title:
        return None  # Skip rows with no title

    applicant = _get("applicant")
    publication_number = _get("publicationNumber")
    application_number = _get("applicationNumber")
    country = _normalize_country(_get("country"))
    publication_date = _normalize_date(_get("publicationDate"))
    legal_status = _normalize_legal_status(_get("legalStatus"))
    abstract = _get("abstract")
    key_claims_summary = _get("keyClaimsSummary")
    inventors_raw = _get("inventors")
    inventors = [inv.strip() for inv in inventors_raw.replace("；", ";").replace("、", ";").replace(",", ";").split(";") if inv.strip()] if inventors_raw else []
    ipc_codes_raw = _get("ipcCodes")
    ipc_codes = [code.strip() for code in ipc_codes_raw.replace("；", ";").replace(",", ";").split(";") if code.strip()] if ipc_codes_raw else []
    url = _get("url")

    return CreatePatentRecordRequest(
        title=title,
        applicant=applicant,
        publication_number=publication_number,
        application_number=application_number,
        country=country,
        publication_date=publication_date or None,
        legal_status=legal_status,
        source="xlsx_import",
        url=url or None,
        abstract=abstract,
        key_claims_summary=key_claims_summary,
        inventors=inventors,
        ipc_codes=ipc_codes,
        notes="Imported from spreadsheet",
    )


def _normalize_country(raw: str) -> str:
    """Normalize country names to ISO codes."""
    COUNTRY_MAP = {  # noqa: N806
        "中国": "CN",
        "美国": "US",
        "欧洲": "EP",
        "日本": "JP",
        "韩国": "KR",
        "香港": "HK",
        "台湾": "TW",
        "英国": "GB",
        "德国": "DE",
        "法国": "FR",
        "加拿大": "CA",
        "澳大利亚": "AU",
        "印度": "IN",
        "新加坡": "SG",
        "世界知识产权": "WO",
        "wipo": "WO",
    }
    norm = raw.lower().strip()
    for key, code in COUNTRY_MAP.items():
        if key in norm:
            return code
    # Already an ISO code
    if len(raw) == 2 and raw.isupper():
        return raw
    return raw


def _normalize_legal_status(raw: str) -> str:
    """Normalize legal status to one of the PatentLegalStatus literals."""
    norm = raw.lower().strip()
    if any(kw in norm for kw in ["授权", "有效", "维持"]):
        return "active"
    if "失效" in norm or "终止" in norm:
        return "lapsed"
    if "公开" in norm or "公布" in norm:
        return "pending"
    if "撤回" in norm or "撤销" in norm:
        return "withdrawn"
    if "届满" in norm or "期满" in norm:
        return "expired"
    if norm in ("active", "granted", "pending", "lapsed", "withdrawn", "expired"):
        return norm
    return "unknown"


def _normalize_date(raw: str) -> str:
    """Normalize date to ISO YYYY-MM-DD format."""
    if not raw:
        return ""
    # Already ISO
    if len(raw) == 10 and raw[4] == "-" and raw[7] == "-":
        return raw
    # Try parsing common formats
    import re
    # YYYY.MM.DD or YYYY/MM/DD
    match = re.match(r"(\d{4})[\./](\d{1,2})[\./](\d{1,2})", raw)
    if match:
        y, m, d = match.groups()
        return f"{y}-{int(m):02d}-{int(d):02d}"
    # YYYYMMDD
    if len(raw) == 8 and raw.isdigit():
        return f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
    return raw  # Return as-is if unparseable


if __name__ == "__main__":
    main()
