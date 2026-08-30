"""Local spreadsheet authoring and editing for Echo."""

from __future__ import annotations

import contextlib
from pathlib import Path
from typing import Any

from runtime.execution.suckers.registry import Skill
from runtime.platform.plugins.bundled._office_io import (
    atomic_package_save,
    create_versioned_backup,
    scoped_path_denial,
)
from runtime.platform.plugins.plugin_base import ModulePlugin

PLUGIN_NAME = "spreadsheets"
_TRUSTED_SOURCE = "plugin://spreadsheets"
_MAX_READ_ROWS = 500
_MAX_READ_COLS = 100

try:  # pragma: no cover - dependency probe
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.cell import range_boundaries

    _OPENPYXL_OK = True
except Exception:  # pragma: no cover
    _OPENPYXL_OK = False


def _require_openpyxl() -> dict[str, Any] | None:
    if _OPENPYXL_OK:
        return None
    return {
        "ok": False,
        "error": "openpyxl 未安装，无法处理 xlsx（安装: pip install openpyxl）",
    }


def _resolve_xlsx(path: Any, *, write: bool = False) -> tuple[Path | None, dict[str, Any] | None]:
    if not isinstance(path, (str, Path)) or not str(path).strip():
        return None, {"ok": False, "error": "path 不能为空"}
    try:
        resolved = Path(str(path)).expanduser().resolve()
    except Exception as exc:
        return None, {"ok": False, "error": f"无效路径: {exc}"}
    if resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None, {"ok": False, "error": "仅支持 .xlsx / .xlsm 文件"}
    denial = scoped_path_denial(resolved, write=write)
    if denial:
        return None, {"ok": False, "error": denial}
    return resolved, None


class SpreadsheetsPlugin(ModulePlugin):
    name = PLUGIN_NAME
    display_name = "Spreadsheets"
    version = "0.1.0"
    description = "本地创建、读取和原位修改 Excel 工作簿，保留公式与样式。"
    author = "Echo"

    def register_skills(self) -> None:
        if self.ctx is None:
            return
        skills = [
            Skill(
                name="spreadsheets.create_xlsx",
                description=(
                    "从结构化 sheets 创建 xlsx。参数:path，sheets:[{name,rows,freeze_panes,"
                    "column_widths,auto_filter,bold_header}]，overwrite 默认 false。"
                ),
                summary="创建 xlsx(path+sheets)",
                affinity=["spreadsheets", "xlsx", "excel", "file", "write", "create"],
                cost_profile="low",
                trusted_source=_TRUSTED_SOURCE,
                handler=self._create_xlsx,
            ),
            Skill(
                name="spreadsheets.read_sheet",
                description=(
                    "读取 xlsx 指定工作表/区域。参数:path，sheet 可选，range 可选(A1:D20)，"
                    "data_only 可选，max_rows/max_cols 有上限。"
                ),
                summary="读取表格区域(path, sheet, range)",
                affinity=["spreadsheets", "xlsx", "excel", "file", "read", "analyze"],
                cost_profile="low",
                trusted_source=_TRUSTED_SOURCE,
                handler=self._read_sheet,
            ),
            Skill(
                name="spreadsheets.update_cells",
                description=(
                    "原位修改 xlsx 单元格并保留其他公式/样式。参数:path，sheet，"
                    "updates:[{cell,value|formula,number_format}]，backup 默认 true。"
                ),
                summary="原位修改 xlsx 单元格",
                affinity=["spreadsheets", "xlsx", "excel", "file", "write", "edit", "formula"],
                cost_profile="low",
                trusted_source=_TRUSTED_SOURCE,
                handler=self._update_cells,
            ),
            Skill(
                name="spreadsheets.workbook_info",
                description="返回 xlsx 工作表、使用区域、合并单元格和公式数量。参数:path。",
                summary="xlsx 工作簿结构统计",
                affinity=["spreadsheets", "xlsx", "excel", "file", "read", "info"],
                cost_profile="low",
                trusted_source=_TRUSTED_SOURCE,
                handler=self._workbook_info,
            ),
        ]
        for skill in skills:
            with contextlib.suppress(Exception):
                self.ctx.register_skill(skill)

    def _create_xlsx(self, **kwargs: Any) -> dict[str, Any]:
        path, err = _resolve_xlsx(kwargs.get("path"), write=True)
        if err:
            return err
        assert path is not None
        if path.exists() and not kwargs.get("overwrite"):
            return {"ok": False, "error": f"目标已存在: {path},需 overwrite=true"}
        dependency_error = _require_openpyxl()
        if dependency_error:
            return dependency_error
        sheets = kwargs.get("sheets")
        if not isinstance(sheets, list) or not sheets:
            return {"ok": False, "error": "sheets 必须是非空数组"}

        workbook = Workbook()
        workbook.remove(workbook.active)
        created: list[dict[str, Any]] = []
        for index, spec in enumerate(sheets):
            if not isinstance(spec, dict):
                return {"ok": False, "error": f"sheets[{index}] 必须是对象"}
            name = str(spec.get("name") or f"Sheet{index + 1}")[:31]
            if name in workbook.sheetnames:
                return {"ok": False, "error": f"工作表名重复: {name}"}
            worksheet = workbook.create_sheet(name)
            rows = spec.get("rows") or []
            if not isinstance(rows, list):
                return {"ok": False, "error": f"{name}.rows 必须是数组"}
            for row in rows:
                worksheet.append(list(row) if isinstance(row, (list, tuple)) else [row])
            if rows and spec.get("bold_header", True):
                fill = PatternFill("solid", fgColor="E8EEF8")
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = fill
            if spec.get("freeze_panes"):
                worksheet.freeze_panes = str(spec["freeze_panes"])
            if spec.get("auto_filter") and worksheet.max_row and worksheet.max_column:
                worksheet.auto_filter.ref = worksheet.dimensions
            widths = spec.get("column_widths") or {}
            if isinstance(widths, dict):
                for column, width in widths.items():
                    with contextlib.suppress(TypeError, ValueError):
                        worksheet.column_dimensions[str(column).upper()].width = min(
                            100, max(1, float(width))
                        )
            created.append({"name": name, "rows": worksheet.max_row, "cols": worksheet.max_column})
        try:
            atomic_package_save(path, workbook.save)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"保存 xlsx 失败: {exc}"}
        return {"ok": True, "path": str(path), "sheets": created}

    def _open(self, path_value: Any, *, data_only: bool = False):
        path, err = _resolve_xlsx(path_value)
        if err:
            return None, None, err
        assert path is not None
        if not path.is_file():
            return None, None, {"ok": False, "error": f"文件不存在: {path}"}
        dependency_error = _require_openpyxl()
        if dependency_error:
            return None, None, dependency_error
        try:
            workbook = load_workbook(
                path,
                data_only=data_only,
                keep_vba=path.suffix.lower() == ".xlsm",
            )
        except Exception as exc:  # noqa: BLE001
            return None, None, {"ok": False, "error": f"无法打开工作簿: {exc}"}
        return path, workbook, None

    @staticmethod
    def _sheet(workbook: Any, requested: Any):
        name = str(requested or workbook.sheetnames[0])
        if name not in workbook.sheetnames:
            return None, {"ok": False, "error": f"工作表不存在: {name}"}
        return workbook[name], None

    def _read_sheet(self, **kwargs: Any) -> dict[str, Any]:
        path, workbook, err = self._open(
            kwargs.get("path"), data_only=bool(kwargs.get("data_only"))
        )
        if err:
            return err
        worksheet, sheet_err = self._sheet(workbook, kwargs.get("sheet"))
        if sheet_err:
            return sheet_err
        try:
            max_rows = min(_MAX_READ_ROWS, max(1, int(kwargs.get("max_rows") or 200)))
            max_cols = min(_MAX_READ_COLS, max(1, int(kwargs.get("max_cols") or 50)))
        except (TypeError, ValueError):
            return {"ok": False, "error": "max_rows / max_cols 必须是整数"}
        min_col = min_row = 1
        max_col = min(worksheet.max_column, max_cols)
        max_row = min(worksheet.max_row, max_rows)
        range_ref = str(kwargs.get("range") or "").strip()
        if range_ref:
            try:
                min_col, min_row, requested_max_col, requested_max_row = range_boundaries(range_ref)
            except ValueError as exc:
                return {"ok": False, "error": f"无效 range: {exc}"}
            max_col = min(requested_max_col, min_col + max_cols - 1)
            max_row = min(requested_max_row, min_row + max_rows - 1)
        rows = [
            [cell.value for cell in row]
            for row in worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
        ]
        return {
            "ok": True,
            "path": str(path),
            "sheet": worksheet.title,
            "range": f"{worksheet.cell(min_row, min_col).coordinate}:{worksheet.cell(max_row, max_col).coordinate}",
            "rows": rows,
            "truncated": worksheet.max_row > max_row or worksheet.max_column > max_col,
            "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
        }

    def _update_cells(self, **kwargs: Any) -> dict[str, Any]:
        path, workbook, err = self._open(kwargs.get("path"))
        if err:
            return err
        assert path is not None
        denial = scoped_path_denial(path, write=True)
        if denial:
            return {"ok": False, "error": denial}
        worksheet, sheet_err = self._sheet(workbook, kwargs.get("sheet"))
        if sheet_err:
            return sheet_err
        updates = kwargs.get("updates")
        if not isinstance(updates, list) or not updates:
            return {"ok": False, "error": "updates 必须是非空数组"}
        changed: list[str] = []
        for index, update in enumerate(updates):
            if not isinstance(update, dict) or not str(update.get("cell") or "").strip():
                return {"ok": False, "error": f"updates[{index}].cell 无效"}
            coordinate = str(update["cell"]).strip().upper()
            try:
                cell = worksheet[coordinate]
            except (KeyError, ValueError) as exc:
                return {"ok": False, "error": f"无效单元格 {coordinate}: {exc}"}
            try:
                if "formula" in update:
                    formula = str(update.get("formula") or "")
                    cell.value = formula if formula.startswith("=") else f"={formula}"
                elif "value" in update:
                    cell.value = update.get("value")
                else:
                    return {"ok": False, "error": f"updates[{index}] 缺少 value 或 formula"}
                if update.get("number_format") is not None:
                    cell.number_format = str(update["number_format"])
            except (AttributeError, TypeError, ValueError) as exc:
                return {
                    "ok": False,
                    "error": f"无法修改单元格 {coordinate}: {exc}。合并区域只能修改左上角单元格",
                }
            changed.append(coordinate)
        backup_path = create_versioned_backup(path) if kwargs.get("backup", True) else None
        try:
            atomic_package_save(path, workbook.save)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"保存 xlsx 失败: {exc}"}
        return {
            "ok": True,
            "path": str(path),
            "sheet": worksheet.title,
            "updated_cells": changed,
            "backup_path": str(backup_path) if backup_path else None,
        }

    def _workbook_info(self, **kwargs: Any) -> dict[str, Any]:
        path, workbook, err = self._open(kwargs.get("path"))
        if err:
            return err
        sheets = []
        for worksheet in workbook.worksheets:
            formulas = sum(
                1
                for row in worksheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            )
            sheets.append(
                {
                    "name": worksheet.title,
                    "rows": worksheet.max_row,
                    "cols": worksheet.max_column,
                    "dimension": worksheet.dimensions,
                    "formula_count": formulas,
                    "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                }
            )
        return {"ok": True, "path": str(path), "sheets": sheets, "size_bytes": path.stat().st_size}


__all__ = ["SpreadsheetsPlugin"]
